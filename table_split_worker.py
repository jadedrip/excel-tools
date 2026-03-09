#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
表格切分工作线程类
功能：在后台执行表格切分操作，避免阻塞UI
作者：jadedrip

"""

from PyQt6.QtCore import QThread, pyqtSignal
import os

class TableSplitWorker(QThread):
    """表格切分工作线程"""
    progress_updated = pyqtSignal(int, int, str)  # 当前进度, 总进度, 当前操作描述
    finished = pyqtSignal(bool, str)  # 是否成功, 结果消息
    file_saved = pyqtSignal(str)  # 文件保存信号
    
    def __init__(self, file_path, sheet_name, header_rows_count, split_columns, output_dir):
        super().__init__()
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.header_rows_count = header_rows_count
        self.split_columns = split_columns
        self.output_dir = output_dir
        self.is_cancelled = False
    
    def run(self):
        """执行表格切分操作"""
        try:
            # 导入必要的模块
            import openpyxl
            from openpyxl.utils import get_column_letter
            from copy import copy
            import os
            
            # 加载Excel文件
            self.progress_updated.emit(0, 100, "正在加载Excel文件...")
            workbook = openpyxl.load_workbook(self.file_path, keep_vba=True, data_only=True)
            
            # 获取当前工作表
            sheet = workbook[self.sheet_name]
            
            # 获取数据范围
            max_row = sheet.max_row
            max_col = sheet.max_column
            
            self.progress_updated.emit(5, 100, f"正在分析数据范围: 行1-{max_row}, 列1-{max_col}")
            
            # 读取保留行数据（保留格式）
            self.progress_updated.emit(10, 100, "正在读取保留行数据...")
            header_rows = []
            for row in sheet.iter_rows(min_row=1, max_row=self.header_rows_count, min_col=1, max_col=max_col):
                header_rows.append(row)
            
            # 按指定列分组数据
            self.progress_updated.emit(20, 100, "正在分组数据...")
            data_groups = {}
            
            # 从保留行之后开始读取数据
            for row_num in range(self.header_rows_count + 1, max_row + 1):
                # 检查是否取消
                if self.is_cancelled:
                    self.finished.emit(False, "操作已取消")
                    return
                
                # 依次检查拆分列，获取分组键
                group_key = None
                filename = None
                
                # 修改：按切分列最后非空行切分
                for col_idx in self.split_columns:
                    if col_idx > max_col:
                        continue  # 跳过超出范围的列
                        
                    cell_value = sheet.cell(row=row_num, column=col_idx).value
                    if cell_value and str(cell_value).strip():
                        group_key = str(cell_value).strip()
                        filename = f"{group_key}.xlsx"
                        # 不break，继续检查后面的列，使用最后一个非空列的值作为分组键
                
                if not group_key:
                    continue  # 跳过所有拆分列都为空的行
                
                # 将行添加到分组
                if group_key not in data_groups:
                    data_groups[group_key] = {
                        "filename": filename,
                        "rows": []
                    }
                
                # 读取整行数据
                row_data = []
                for col_num in range(1, max_col + 1):
                    cell = sheet.cell(row=row_num, column=col_num)
                    row_data.append(cell)
                
                data_groups[group_key]["rows"].append(row_data)
            
            if not data_groups:
                self.finished.emit(False, "没有找到可分组的数据")
                return
            
            # 为每个分组创建新的Excel文件
            total_groups = len(data_groups)
            self.progress_updated.emit(30, 100, f"开始创建 {total_groups} 个Excel文件...")
            
            for i, (group_key, group_data) in enumerate(data_groups.items()):
                # 检查是否取消
                if self.is_cancelled:
                    self.finished.emit(False, "操作已取消")
                    return
                
                # 更新进度
                progress = 30 + int((i / total_groups) * 60)
                self.progress_updated.emit(progress, 100, f"正在处理文件 {i+1}/{total_groups}: {group_key}")
                
                # 创建新的工作簿
                new_workbook = openpyxl.Workbook()
                new_sheet = new_workbook.active
                new_sheet.title = self.sheet_name
                
                # 复制保留行（包括格式和合并单元格）
                for row_idx, row in enumerate(header_rows, 1):
                    for col_idx, cell in enumerate(row, 1):
                        # 复制单元格值
                        new_cell = new_sheet.cell(row=row_idx, column=col_idx, value=cell.value)
                        
                        # 复制单元格格式
                        if cell.has_style:
                            try:
                                new_cell.font = copy(cell.font)
                                new_cell.border = copy(cell.border)
                                new_cell.fill = copy(cell.fill)
                                new_cell.number_format = cell.number_format
                                new_cell.protection = copy(cell.protection)
                                new_cell.alignment = copy(cell.alignment)
                            except Exception:
                                # 忽略格式复制错误，继续执行
                                pass
                
                # 复制合并单元格
                for merged_cell in sheet.merged_cells.ranges:
                    # 只复制保留行内的合并单元格
                    if merged_cell.min_row <= self.header_rows_count:
                        try:
                            new_sheet.merge_cells(str(merged_cell))
                        except Exception:
                            # 忽略合并单元格错误，继续执行
                            pass
                
                # 添加数据行
                for row_idx, row in enumerate(group_data["rows"], self.header_rows_count + 1):
                    for col_idx, cell in enumerate(row, 1):
                        # 复制单元格值
                        new_cell = new_sheet.cell(row=row_idx, column=col_idx, value=cell.value)
                        
                        # 复制单元格格式
                        if cell.has_style:
                            try:
                                new_cell.font = copy(cell.font)
                                new_cell.border = copy(cell.border)
                                new_cell.fill = copy(cell.fill)
                                new_cell.number_format = cell.number_format
                                new_cell.protection = copy(cell.protection)
                                new_cell.alignment = copy(cell.alignment)
                            except Exception:
                                # 忽略格式复制错误，继续执行
                                pass
                
                # 复制列宽
                for col_idx in range(1, max_col + 1):
                    col_letter = get_column_letter(col_idx)
                    if col_letter in sheet.column_dimensions:
                        try:
                            new_sheet.column_dimensions[col_letter].width = sheet.column_dimensions[col_letter].width
                        except Exception:
                            # 忽略列宽复制错误，继续执行
                            pass
                
                # 保存文件
                output_path = os.path.join(self.output_dir, group_data["filename"])
                try:
                    new_workbook.save(output_path)
                    self.file_saved.emit(output_path)
                except Exception as e:
                    # 继续处理其他文件
                    pass
            
            self.progress_updated.emit(100, 100, "表格切分完成")
            self.finished.emit(True, f"表格切分完成！共生成 {len(data_groups)} 个文件\n输出目录: {os.path.abspath(self.output_dir)}")
            
        except Exception as e:
            self.finished.emit(False, f"表格切分失败: {str(e)}")
    
    def cancel(self):
        """取消操作"""
        self.is_cancelled = True