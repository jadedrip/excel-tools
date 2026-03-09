#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
通用Excel拆分工具

功能：
1. 支持命令行指定输入文件、工作表、保留行数
2. 支持指定多个拆分列，依次判断
3. 输出目录使用工作表名
4. 保留原格式和合并单元格
5. 兼容多种Excel格式

使用示例：
python excel_splitter_generic.py -f "附件1：自有房地资源数据收集=生产类.xlsx" -s "房产数据收集模板" -r 8 -c 4 5
python excel_splitter_generic.py -f "附件1：自有房地资源数据收集=生产类.xlsx" -s "土地数据收集模板" -r 7 -c 6 7
"""

import os
import argparse
import openpyxl
from openpyxl.utils import get_column_letter
from copy import copy


def col_letter_to_index(letter):
    """将Excel列字母转换为数字索引（从1开始）
    
    Args:
        letter: Excel列字母，如'A', 'B', 'AA'等
        
    Returns:
        int: 列索引，从1开始
    """
    letter = letter.upper()
    index = 0
    for char in letter:
        index = index * 26 + (ord(char) - ord('A') + 1)
    return index


def parse_args():
    """解析命令行参数"""
    parser = argparse.ArgumentParser(description='通用Excel拆分工具')
    parser.add_argument('-f', '--file', type=str, required=True, help='输入Excel文件路径')
    parser.add_argument('-s', '--sheet', type=str, default=None, help='要处理的工作表名称')
    parser.add_argument('-i', '--sheet-index', type=int, default=None, help='要处理的工作表索引（从1开始），1代表第一个sheet，优先级高于工作表名称')
    parser.add_argument('-r', '--rows', type=int, required=True, help='要保留的前几行数量')
    parser.add_argument('-c', '--columns', type=str, nargs='+', required=True, help='拆分列的字母索引（如A、B、C、E），依次判断')
    
    return parser.parse_args()


def main():
    """主函数"""
    args = parse_args()
    
    input_file = args.file
    sheet_name = args.sheet
    header_rows_count = args.rows
    
    # 将列字母转换为数字索引
    split_columns = []
    for col_letter in args.columns:
        try:
            col_index = col_letter_to_index(col_letter)
            split_columns.append(col_index)
        except Exception as e:
            print(f"错误：无效的列字母 '{col_letter}'")
            return
    
    # 验证输入文件
    if not os.path.exists(input_file):
        print(f"错误：输入文件 '{input_file}' 不存在")
        return
    
    print(f"输入文件: {input_file}")
    print(f"保留行数: {header_rows_count}")
    print(f"拆分列索引: {split_columns}")
    
    # 加载Excel文件
    try:
        workbook = openpyxl.load_workbook(input_file, keep_vba=True, data_only=True)
    except Exception as e:
        print(f"错误：无法加载Excel文件 - {e}")
        return
    
    # 获取所有工作表名称
    sheet_names = workbook.sheetnames
    selected_sheet_name = None
    
    # 确定要处理的工作表
    if args.sheet_index is not None:
        # 使用工作表索引（从1开始），转换为0-based索引
        real_index = args.sheet_index - 1
        if 0 <= real_index < len(sheet_names):
            selected_sheet_name = sheet_names[real_index]
            print(f"使用工作表索引 {args.sheet_index}，对应工作表：{selected_sheet_name}")
        else:
            print(f"错误：工作表索引 {args.sheet_index} 超出范围（有效范围：1-{len(sheet_names)}）")
            print(f"可用工作表：{sheet_names}")
            return
    elif args.sheet is not None:
        # 使用工作表名称
        if args.sheet in sheet_names:
            selected_sheet_name = args.sheet
        else:
            print(f"错误：工作表 '{args.sheet}' 不存在")
            print(f"可用工作表：{sheet_names}")
            return
    else:
        # 默认处理第一个工作表
        selected_sheet_name = sheet_names[0]
        print(f"未指定工作表，默认处理第一个工作表：{selected_sheet_name}")
    
    # 更新sheet_name变量
    sheet_name = selected_sheet_name
    
    # 输出目录使用工作表名
    output_dir = sheet_name
    if not os.path.exists(output_dir):
        os.makedirs(output_dir, exist_ok=True)
    
    print(f"处理工作表: {sheet_name}")
    print(f"输出目录: {output_dir}")
    
    sheet = workbook[sheet_name]
    
    # 获取数据范围
    max_row = sheet.max_row
    max_col = sheet.max_column
    
    print(f"数据范围: 行1-{max_row}, 列1-{max_col}")
    
    # 读取保留行数据（保留格式）
    header_rows = []
    for row in sheet.iter_rows(min_row=1, max_row=header_rows_count, min_col=1, max_col=max_col):
        header_rows.append(row)
    
    # 按指定列分组数据
    data_groups = {}
    
    # 从保留行之后开始读取数据
    for row_num in range(header_rows_count + 1, max_row + 1):
        # 依次检查拆分列，获取分组键
        group_key = None
        filename = None
        
        for col_idx in split_columns:
            if col_idx > max_col:
                continue  # 跳过超出范围的列
                
            cell_value = sheet.cell(row=row_num, column=col_idx).value
            if cell_value and str(cell_value).strip():
                group_key = str(cell_value).strip()
                filename = f"{group_key}.xlsx"
                break  # 找到第一个非空列，使用该值作为分组键
        
        if not group_key:
            continue  # 跳过所有拆分列都为空的行
        
        # 将行添加到分组
        if group_key not in data_groups:
            data_groups[group_key] = {
                "filename": filename,
                "rows": []
            }
        
        # 读取整行数据
        row_data = []
        for col_num in range(1, max_col + 1):
            cell = sheet.cell(row=row_num, column=col_num)
            row_data.append(cell)
        
        data_groups[group_key]["rows"].append(row_data)
    
    print(f"共分为 {len(data_groups)} 个分组")
    
    # 为每个分组创建新的Excel文件
    for i, (group_key, group_data) in enumerate(data_groups.items(), 1):
        print(f"\n处理分组 {i}/{len(data_groups)}: {group_key}")
        
        # 创建新的工作簿
        new_workbook = openpyxl.Workbook()
        new_sheet = new_workbook.active
        new_sheet.title = sheet_name
        
        # 复制保留行（包括格式和合并单元格）
        print(f"  复制前 {header_rows_count} 行格式...")
        for row_idx, row in enumerate(header_rows, 1):
            for col_idx, cell in enumerate(row, 1):
                # 复制单元格值
                new_cell = new_sheet.cell(row=row_idx, column=col_idx, value=cell.value)
                
                # 复制单元格格式
                if cell.has_style:
                    try:
                        new_cell.font = copy(cell.font)
                        new_cell.border = copy(cell.border)
                        new_cell.fill = copy(cell.fill)
                        new_cell.number_format = cell.number_format
                        new_cell.protection = copy(cell.protection)
                        new_cell.alignment = copy(cell.alignment)
                    except Exception as e:
                        # 忽略格式复制错误，继续执行
                        pass
        
        # 复制合并单元格
        print(f"  复制合并单元格...")
        for merged_cell in sheet.merged_cells.ranges:
            # 只复制保留行内的合并单元格
            if merged_cell.min_row <= header_rows_count:
                try:
                    new_sheet.merge_cells(str(merged_cell))
                except Exception as e:
                    # 忽略合并单元格错误，继续执行
                    pass
        
        # 添加数据行
        print(f"  添加 {len(group_data['rows'])} 行数据...")
        for row_idx, row in enumerate(group_data["rows"], header_rows_count + 1):
            for col_idx, cell in enumerate(row, 1):
                # 复制单元格值
                new_cell = new_sheet.cell(row=row_idx, column=col_idx, value=cell.value)
                
                # 复制单元格格式
                if cell.has_style:
                    try:
                        new_cell.font = copy(cell.font)
                        new_cell.border = copy(cell.border)
                        new_cell.fill = copy(cell.fill)
                        new_cell.number_format = cell.number_format
                        new_cell.protection = copy(cell.protection)
                        new_cell.alignment = copy(cell.alignment)
                    except Exception as e:
                        # 忽略格式复制错误，继续执行
                        pass
        
        # 复制列宽
        print(f"  复制列宽...")
        for col_idx in range(1, max_col + 1):
            col_letter = get_column_letter(col_idx)
            if col_letter in sheet.column_dimensions:
                try:
                    new_sheet.column_dimensions[col_letter].width = sheet.column_dimensions[col_letter].width
                except Exception as e:
                    # 忽略列宽复制错误，继续执行
                    pass
        
        # 保存文件
        output_path = os.path.join(output_dir, group_data["filename"])
        try:
            new_workbook.save(output_path)
            print(f"  已保存文件: {output_path}")
        except Exception as e:
            print(f"  错误：保存文件失败 - {e}")
            continue
    
    print(f"\n处理完成！共生成 {len(data_groups)} 个文件")
    print(f"输出目录: {os.path.abspath(output_dir)}")


if __name__ == "__main__":
    main()
