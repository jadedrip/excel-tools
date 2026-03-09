#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
修复Excel文件中的styles.xml问题
问题：fills元素中存在空的自闭合fill标签
"""

import os
import zipfile
import re
import shutil


def backup_file(file_path):
    """备份文件"""
    backup_path = file_path.replace('.xlsx', '_backup.xlsx')
    if not os.path.exists(backup_path):
        shutil.copy2(file_path, backup_path)
        print(f"[信息] 已备份原文件到: {backup_path}")
    return backup_path


def extract_and_fix_styles(file_path):
    """提取并修复styles.xml"""
    print("=" * 60)
    print("提取并分析styles.xml")
    print("=" * 60)
    
    with zipfile.ZipFile(file_path, 'r') as zip_file:
        if 'xl/styles.xml' not in zip_file.namelist():
            print("[错误] styles.xml文件不存在")
            return False
        
        content = zip_file.read('xl/styles.xml').decode('utf-8')
    
    print(f"[信息] 原styles.xml大小: {len(content)} bytes")
    
    # 查找问题：空fill标签
    # 模式：<fill/> 或 <fill />
    empty_fill_pattern = re.compile(r'<fills[^>]*>.*?<fill\s*/>.*?</fills>', re.DOTALL | re.IGNORECASE)
    
    matches = empty_fill_pattern.findall(content)
    print(f"[调试] 找到 {len(matches)} 个可能的空fill模式")
    
    # 查找所有的fill标签
    fills_pattern = re.compile(r'<fills[^>]*>(.*?)</fills>', re.DOTALL | re.IGNORECASE)
    fills_matches = fills_pattern.findall(content)
    
    if fills_matches:
        fills_content = fills_matches[0]
        print(f"[信息] fills元素内容长度: {len(fills_content)}")
        
        # 查找空的fill标签
        empty_fill_count = len(re.findall(r'<fill\s*/>', fills_content))
        print(f"[信息] 找到 {empty_fill_count} 个空fill标签")
        
        # 打印fill元素列表
        fill_items = re.findall(r'<fill[^>]*>.*?</fill>', fills_content, re.DOTALL | re.IGNORECASE)
        print(f"[信息] 总共 {len(fill_items)} 个fill元素")
        
        for i, fill in enumerate(fill_items):
            # 清理空白字符进行显示
            fill_clean = fill.strip().replace('\n', '').replace('  ', '')
            if len(fill_clean) > 60:
                fill_clean = fill_clean[:60] + '...'
            print(f"  [{i+1:2d}] {fill_clean}")
    
    # 修复：移除空的fill标签
    fixed_content = re.sub(r'<fill\s*/>', '', content)
    
    # 验证修复
    empty_fill_remaining = len(re.findall(r'<fill\s*/>', fixed_content))
    print(f"[信息] 修复后空fill标签数量: {empty_fill_remaining}")
    
    if empty_fill_remaining == 0:
        print("[成功] 已移除所有空fill标签")
    
    return fixed_content


def write_fixed_styles(file_path, fixed_content):
    """将修复后的styles.xml写回文件"""
    print("\n" + "=" * 60)
    print("写回修复后的文件")
    print("=" * 60)
    
    # 备份原文件
    backup_path = backup_file(file_path)
    
    # 创建临时目录
    temp_dir = "temp_xlsx_extract"
    if os.path.exists(temp_dir):
        shutil.rmtree(temp_dir)
    os.makedirs(temp_dir)
    
    try:
        # 解压原文件
        with zipfile.ZipFile(file_path, 'r') as zip_file:
            zip_file.extractall(temp_dir)
        
        # 替换styles.xml
        styles_path = os.path.join(temp_dir, 'xl', 'styles.xml')
        with open(styles_path, 'w', encoding='utf-8') as f:
            f.write(fixed_content)
        print(f"[信息] 已写入修复后的styles.xml")
        
        # 重新压缩为xlsx文件
        new_file_path = file_path.replace('.xlsx', '_fixed.xlsx')
        if os.path.exists(new_file_path):
            os.remove(new_file_path)
        
        with zipfile.ZipFile(new_file_path, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for root, dirs, files in os.walk(temp_dir):
                for file in files:
                    file_path_abs = os.path.join(root, file)
                    arcname = os.path.relpath(file_path_abs, temp_dir)
                    zip_file.write(file_path_abs, arcname)
        
        print(f"[成功] 已创建修复后的文件: {new_file_path}")
        print(f"[信息] 文件大小: {os.path.getsize(new_file_path)} bytes")
        
        return new_file_path
        
    finally:
        # 清理临时目录
        if os.path.exists(temp_dir):
            shutil.rmtree(temp_dir)


def test_fixed_file(file_path):
    """测试修复后的文件"""
    print("\n" + "=" * 60)
    print("测试修复后的文件")
    print("=" * 60)
    
    try:
        import openpyxl
        
        print(f"[信息] 尝试打开: {file_path}")
        workbook = openpyxl.load_workbook(file_path, keep_vba=True, data_only=True)
        print(f"[成功] 文件打开成功!")
        print(f"[信息] 工作表列表: {workbook.sheetnames}")
        
        # 读取第一个工作表的数据
        for sheet_name in workbook.sheetnames[:2]:
            sheet = workbook[sheet_name]
            print(f"\n[信息] 工作表 '{sheet_name}':")
            print(f"  - 最大行数: {sheet.max_row}")
            print(f"  - 最大列数: {sheet.max_column}")
            
            # 读取前几行
            for row_idx, row in enumerate(sheet.iter_rows(max_row=3, max_col=5), 1):
                row_data = []
                for cell in row:
                    val = cell.value
                    if val is not None:
                        val_str = str(val)[:15]
                    else:
                        val_str = "None"
                    row_data.append(val_str)
                print(f"    行{row_idx}: {row_data}")
        
        return True
        
    except Exception as e:
        print(f"[错误] 测试失败: {e}")
        import traceback
        traceback.print_exc()
        return False


def create_workaround_reader():
    """创建一个使用其他方法的读取器作为备选方案"""
    print("\n" + "=" * 60)
    print("创建备选读取方案")
    print("=" * 60)
    
    # 检查是否安装了xlrd或其他库
    try:
        import xlrd
        print("[信息] xlrd可用")
    except ImportError:
        print("[信息] xlrd不可用")
    
    try:
        import pandas as pd
        print(f"[信息] pandas版本: {pd.__version__}")
    except ImportError:
        print("[信息] pandas不可用")


def main():
    """主函数"""
    file_path = "子站设备调试设备清单.xlsx"
    
    if not os.path.exists(file_path):
        print(f"[错误] 文件不存在: {file_path}")
        print("请确保测试文件在当前目录下")
        return
    
    print("Excel文件样式问题修复工具")
    print("=" * 60)
    
    # 步骤1: 分析并修复styles.xml
    fixed_content = extract_and_fix_styles(file_path)
    
    if not fixed_content:
        return
    
    # 步骤2: 写回修复后的文件
    fixed_file = write_fixed_styles(file_path, fixed_content)
    
    # 步骤3: 测试修复后的文件
    if os.path.exists(fixed_file):
        success = test_fixed_file(fixed_file)
        
        if success:
            print("\n" + "=" * 60)
            print("修复成功!")
            print("=" * 60)
            print(f"修复后的文件: {fixed_file}")
            print("\n建议:")
            print("1. 使用修复后的文件进行测试")
            print("2. 如果需要永久修复，请用Excel重新保存原文件")
        else:
            print("\n[警告] 修复后的文件仍有问题")
    else:
        print("[错误] 创建修复后的文件失败")


if __name__ == "__main__":
    main()
